# coding=utf-8
import numpy as np
import SQL
import matplotlib as mpl
mpl.use('Agg')
import matplotlib.pyplot as plt
import re
from prettytable import PrettyTable
from PIL import Image, ImageDraw, ImageFont
import openpyxl
import datetime

def yield_type(btime, etime, work_type, fname,NickName):
    lx = {
        '收货': [],
        '上架': [],
        '拣货': [],
        '包装': [],
        '盘点': [],
        '移库': []
    }

    if work_type not in lx:
        return "操作类型不存在！目前可供查询：收货,上架,拣货,包装,盘点,移库"

    db = SQL.WMS(NickName)

    data = db.yield_table(btime, etime)

    if data =='执行结果出错':
        return "数据库查询结果出错，请修正参数"

    user = []
    for k, v in data.items():
        name = re.findall(re.compile(r'[(](.*?)[)]', re.S), k)
        user.append(name[0])
        lx['收货'].append(v[0])
        lx['上架'].append(v[1])
        lx['拣货'].append(v[2])
        lx['包装'].append(v[3])
        lx['盘点'].append(v[4])
        lx['移库'].append(v[5])

    sj = lx[work_type]

    user_list = []  # 图表姓名
    cl = []  # 对应产量

    for i, j in enumerate(sj):
        if j:
            user_list.append(user[i])
            cl.append(j)
    sum_cl = sum(cl)
    sum_cl = work_type + "总产量:" + str(sum_cl)
    loan_grade = cl
    # 图表字体为华文细黑，字号为15
    plt.rc('font', family='SIMHEI', size=20)

    # 创建柱状图，数据源x,y来源，设置颜色，透明度和外边框颜色
    plt.bar(user_list, loan_grade, color='#99CC01', alpha=0.8, align='center', edgecolor='white')
    # 设置x轴标签
    # plt.xlabel('账号')
    # 设置y周标签
    # plt.ylabel('产量')
    # 设置图表标题
    plt.title("{}——{}/{}产量表         打印时间：{}".format(btime, etime, work_type,fname))
    # 设置图例的文字和在图表中的位置
    plt.legend([sum_cl], loc='upper right')

    # 设置背景网格线的颜色，样式，尺寸和透明度
    plt.grid(color='#95a5a6', linestyle='--', linewidth=1, axis='y', alpha=0.4)
    plt.xticks(rotation=45)  # 轴标签倾斜度数

    # 循环，为每个柱形添加文本标注
    # 居中对齐
    for x, y in zip(user_list, loan_grade):
        plt.text(x, y + 0.1, str(y), ha='center')

    # 显示图表
    # plt.show()

    fig = plt.gcf()
    fig.set_size_inches(18.5, 10.5)  # 设置保存大小

    plt.savefig(fname, dpi=200)  # 保存，dip参数设置分辨率
    plt.close('all')  # 关闭图表释放内存



def chayi(row):

    tab = PrettyTable()
    # 设置表头
    tab.field_names = row[0]
    # 表格内容插入
    for i in row[1:]:
        tab.add_row(i)

    tab_info = str(tab)
    space = 5

    # PIL模块中，确定写入到图片中的文本字体
    # font = ImageFont.truetype( 15, encoding='utf-8')
    font = ImageFont.truetype('/usr/share/fonts/deepin-font-install/SimHei/SIMHEI.TTF', 24, encoding='utf-8')
    # Image模块创建一个图片对象
    im = Image.new('RGB', (10, 10), (255, 255, 255))
    # ImageDraw向图片中进行操作，写入文字或者插入线条都可以
    draw = ImageDraw.Draw(im, "RGB")
    # 根据插入图片中的文字内容和字体信息，来确定图片的最终大小
    img_size = draw.multiline_textsize(tab_info, font=font)  # 设置字体才能显示中文
    # 图片初始化的大小为10-10，现在根据图片内容要重新设置图片的大小
    im_new = im.resize((img_size[0] + space * 2, img_size[1] + space * 2))
    del draw
    del im
    draw = ImageDraw.Draw(im_new, 'RGB')
    # 批量写入到图片中，这里的multiline_text会自动识别换行符
    draw.multiline_text((space, space), tab_info, fill=(0, 0, 0), font=font)

    im_new.save('CHAYI.PNG', "PNG")
    del draw




def bf(NickName,file):
    db = SQL.WMS(NickName)
    data = db.copybf()


    filename = file +'.xlsx'

    wb = openpyxl.Workbook()  # 打开文件
    sheet = wb.active  # 激活sheet表格
    sheet.title = "华南1仓库存备份"  # 添加sheet表格名称

    sheet.cell(row=1, column=1, value='仓库代码')
    sheet.cell(row=1, column=2, value='供应商编号')
    sheet.cell(row=1, column=3, value='货位')
    sheet.cell(row=1, column=4, value='条码')
    sheet.cell(row=1, column=5, value='描述')
    sheet.cell(row=1, column=6, value='在库数量')
    sheet.cell(row=1, column=7, value='移入数量')
    sheet.cell(row=1, column=8, value='销售渠道')
    sheet.cell(row=1, column=9, value='分配数量')
    sheet.cell(row=1, column=10, value='状态')

    ind = 2
    for i in data:
        sheet.cell(row=ind, column=1, value=i[0])
        sheet.cell(row=ind, column=2, value=i[1])
        sheet.cell(row=ind, column=3, value=i[2])
        sheet.cell(row=ind, column=4, value=i[3])
        sheet.cell(row=ind, column=5, value=i[4])
        sheet.cell(row=ind, column=6, value=i[5])
        sheet.cell(row=ind, column=7, value=i[6])
        sheet.cell(row=ind, column=8, value=i[7])
        sheet.cell(row=ind, column=9, value=i[8])
        sheet.cell(row=ind, column=10, value=i[9])
        ind += 1

    wb.save('backups/'+filename)




if __name__ == '__main__':
    # yield_type('2018-9-6', '2018-9-7', '上架', 'qq','搬仓管理群')

    bf()
